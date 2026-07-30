import openpyxl

wb = openpyxl.load_workbook('Supreme Ind_Format.xlsx')
ws = wb.active

def get_color(cell):
    fill = cell.fill
    if fill.patternType is None:
        return "None"
    if fill.fgColor.type == 'rgb':
        return fill.fgColor.rgb
    elif fill.fgColor.type == 'theme':
        return f"theme:{fill.fgColor.theme} tint:{fill.fgColor.tint}"
    return str(fill.fgColor.type)

def get_font_color(cell):
    color = cell.font.color
    if not color:
        return "None"
    if color.type == 'rgb':
        return color.rgb
    elif color.type == 'theme':
        return f"theme:{color.theme} tint:{color.tint}"
    return str(color.type)

samples = [
    ("B1 (Company)", ws['B1']),
    ("C2 (Period)", ws['C2']),
    ("B3 (Gross Sales)", ws['B3']),
    ("B4 (Growth)", ws['B4']),
    ("B13 (Raw Mat)", ws['B13']),
    ("B28 (EBITDA)", ws['B28'])
]

for name, cell in samples:
    print(f"{name}: Fill: {get_color(cell)}, Font: {get_font_color(cell)}, Bold: {cell.font.bold}")

print("GridLines property in XML:", ws.sheet_view.showGridLines)
