"""
excel_writer.py  –  Generates the Supreme-format .xlsx from parsed data.

Strategy
--------
1. Load the template Supreme Ind_Format.xlsx as a reference for row labels
   and formatting.
2. Create a fresh workbook with the same sheet structure.
3. Write headers (company name, period columns).
4. Write data rows using the ROW_SPEC mapping.
5. Write Excel formulas for derived rows.
6. Apply basic formatting (bold headers, number formats).
"""

import re
from pathlib import Path
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xlnumbers)
from openpyxl.utils import get_column_letter, column_index_from_string
from parser import get_sorted_periods

# ---------------------------------------------------------------------------
# ROW SPEC  –  maps row_index → (label, source_metric_or_formula, category)
# category: 'PL' | 'BS' | 'CF' | 'DERIVED' | 'HEADER' | 'BLANK' | 'RATIO'
# ---------------------------------------------------------------------------

ROW_SPEC = {
    1:  ('header_company',  None,                    'HEADER'),
    2:  ('header_periods',  None,                    'HEADER'),
    3:  ('Gross Sales',              'gross_sales',   'PL'),
    4:  ('Growth (% YOY)',           'gs_growth',     'DERIVED'),
    5:  ('Excise Duty',              'excise_duty',   'PL'),
    6:  ('Net Sales',                'net_sales',     'DERIVED'),   # formula: =row3-row5
    7:  ('Growth (% YOY)',           'ns_growth',     'DERIVED'),
    8:  ('Other Operating Income',   'other_op_income','PL'),
    9:  ('Total Income',             'total_income',  'DERIVED'),   # formula: =row6+row8
    10: ('Growth (% YOY)',           'ti_growth',     'DERIVED'),
    11: ('',                         None,            'BLANK'),
    12: ('',                         None,            'BLANK'),
    13: ('Raw Material Consumed',    'raw_material',  'PL'),
    14: ('% sales',                  'rm_pct',        'DERIVED'),
    15: ('Stock Adjustment',         'stock_adj',     'PL'),
    16: ('% sales',                  'sa_pct',        'DERIVED'),
    17: ('Purchase of Finished Goods','purchase_fg',  'PL'),
    18: ('% sales',                  'pfg_pct',       'DERIVED'),
    19: ('Employee Expenses',        'employee_exp',  'PL'),
    20: ('% sales',                  'emp_pct',       'DERIVED'),
    21: ('Other Expenses',           'other_exp',     'PL'),
    22: ('% sales',                  'oe_pct',        'DERIVED'),
    23: ('COGS',                     'cogs',          'DERIVED'),
    24: ('Gross Profit (GP)',         'gross_profit',  'DERIVED'),
    25: ('GP Margin (%)',             'gp_margin',     'DERIVED'),
    26: ('Total Expenditure',        'total_exp_formula', 'DERIVED'),  # formula: sum of cost lines
    27: ('% sales',                  'te_pct',        'DERIVED'),
    28: ('PBIDT (EBITDA)',            'ebitda',        'DERIVED'),
    29: ('EBITDA Growth (% YOY)',     'ebitda_growth', 'DERIVED'),
    30: ('EBITDA Margin (%)',         'ebitda_margin', 'DERIVED'),
    31: ('PBIT',                     'pbit',          'DERIVED'),
    32: ('PBIT Growth (% YOY)',       'pbit_growth',   'DERIVED'),
    33: ('PBIT Margin (%)',           'ebit_margin',   'DERIVED'),
    34: ('',                         None,            'BLANK'),
    35: ('Other Income',             'other_income',  'PL'),
    36: ('Interest',                 'interest',      'PL'),
    37: ('PBDT',                     'pbdt',          'DERIVED'),
    38: ('Depreciation',             'depreciation',  'PL'),
    39: ('PBT',                      'pbt',           'DERIVED'),
    40: ('PBT Growth (% YOY)',        'pbt_growth',    'DERIVED'),
    41: ('Tax (Current)',            'tax_current',   'PL'),
    42: ('Fringe Benefit Tax',       None,            'BLANK'),
    43: ('Deferred Tax',             'deferred_tax',  'PL'),
    44: ('Reported PAT',             'pat',           'DERIVED'),
    45: ('Minority Interest',        'minority_interest','PL'),
    46: ('P/L of Associate Company', None,            'BLANK'),
    47: ('Net Profit after MI',      'net_profit_mi', 'PL'),
    48: ('Extra-ordinary Items',     'exceptional_items','PL'),
    49: ('Adjusted PAT',             'pat',           'PL'),
    50: ('PAT Growth (% YOY)',        'pat_growth',    'DERIVED'),
    51: ('Equity (Share Capital)',    'equity',        'BS'),
    52: ('Reserve & Surplus',        'reserves',      'BS'),
    53: ('Debt',                     'debt',          'BS'),
    54: ('Cash Equivalents',         'cash',          'BS'),
    55: ('Net Block',                'net_block',     'BS'),
    56: ('Book Value (Rs.)',          'book_value_ps', 'DERIVED'),
    57: ('EBITDA Margin (%)',         'ebitda_margin', 'RATIO'),
    58: ('Tax/PBT (%)',               'tax_pbt_pct',   'RATIO'),
    59: ('EPS (Rs.)',                 'eps',           'RATIO'),
    60: ('D:E Ratio',                'de_ratio',      'RATIO'),
    61: ('ROE (%)',                   'roe',           'RATIO'),
    62: ('ROCE (%)',                  'roce',          'RATIO'),
    63: ('Face Value (Rs.)',          'face_value',    'BS'),
    64: ('No. of Shares (Cr.)',       'num_shares',    'DERIVED'),
    65: ('CMP (Rs.)',                 None,            'BLANK'),
    66: ('Market Cap (Rs. Cr.)',      None,            'BLANK'),
    67: ('PE (x)',                    None,            'BLANK'),
    68: ('Target Multiple (x)',       None,            'BLANK'),
    69: ('Target Price (Rs.)',        None,            'BLANK'),
    70: ('Upside (%)',                None,            'BLANK'),
    71: ('',                         None,            'BLANK'),
    72: ('Working Capital',          'working_capital','BS'),
    73: ('OCF',                      'ocf',           'CF'),
    74: ('Capex',                    'capex',         'CF'),
    75: ('Dividend',                 'dividend',      'CF'),
    76: ('NOPAT',                    'nopat',         'DERIVED'),
    77: ('Capital Employed',         'cap_employed',  'DERIVED'),
    78: ('Invested Capital',         'invested_cap',  'DERIVED'),
    79: ('ROIC (%)',                  'roic',          'DERIVED'),
}

# ---------------------------------------------------------------------------
# Colour palette (Supreme Ind Format)
# ---------------------------------------------------------------------------
COL_HEADER_FILL   = 'C0C0C0'   # Silver
GROWTH_FILL       = 'FFFF99'   # Pale Yellow
WHITE             = 'FFFFFF'
BLACK             = '000000'

def _font(bold=False, color=BLACK, size=9):
    return Font(bold=bold, color=color, name='Calibri', size=size)

def _fill(hex_color):
    if hex_color == WHITE:
        return PatternFill(fill_type=None) # No fill is default for white background
    return PatternFill('solid', fgColor=hex_color)

def _border():
    s = Side(border_style='thin', color='000000') # Black borders (grid lines)
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _right():
    return Alignment(horizontal='right', vertical='center')


# ---------------------------------------------------------------------------
# Quarterly / Annual column ordering helpers
# ---------------------------------------------------------------------------

def _sort_key(p):
    fy_re  = re.compile(r'^(\d)QFY(\d{2})$')
    ann_re = re.compile(r'^FY(\d{2})$')
    m = fy_re.match(p)
    if m:
        return (int(m.group(2)), int(m.group(1)), 0)
    m = ann_re.match(p)
    if m:
        return (int(m.group(1)), 5, 1)
    return (99, 9, 9)


def _is_annual(p):
    return bool(re.match(r'^FY\d{2}$', p))


def _is_quarterly(p):
    return bool(re.match(r'^\dQFY\d{2}$', p))


# ---------------------------------------------------------------------------
# Main writer
# ---------------------------------------------------------------------------

def write_excel(data: dict, company_name: str, output_path: str,
                template_path: str = None):
    """
    Build and save the formatted Excel workbook.

    Parameters
    ----------
    data         : parsed financial dict from parser.parse_all()
    company_name : string displayed in B1
    output_path  : where to save the .xlsx
    template_path: optional path to original template (for style reference)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Sanitize sheet title (max 31 chars, no special chars)
    sanitized_title = re.sub(r'[*?:/\\\[\]]', '', company_name)[:30].strip() or 'Financials'
    ws.title = sanitized_title
    ws.sheet_view.showGridLines = False

    # -----------------------------------------------------------------------
    # 1. Build period columns (sorted, deduplicated)
    # -----------------------------------------------------------------------
    # Group periods chronologically: 4 quarters followed by their corresponding annual FY column
    fy_groups = {}
    fy_re = re.compile(r'^(\d)QFY(\d{2})$')
    ann_re = re.compile(r'^FY(\d{2})$')

    for p in data.keys():
        mq = fy_re.match(p)
        ma = ann_re.match(p)
        if mq:
            fy = int(mq.group(2))
            fy_groups.setdefault(fy, {'q': [], 'a': None})['q'].append(p)
        elif ma:
            fy = int(ma.group(1))
            fy_groups.setdefault(fy, {'q': [], 'a': None})['a'] = p

    ordered_periods = []
    for fy in sorted(fy_groups.keys()):
        # Sort quarters for this FY (1Q to 4Q)
        quarters = sorted(fy_groups[fy]['q'], key=_sort_key)
        ordered_periods.extend(quarters)
        if fy_groups[fy]['a']:
            ordered_periods.append(fy_groups[fy]['a'])

    # Column assignments: col A = row-category, col B = label, col C onwards = periods
    LABEL_COL   = 2   # B
    FIRST_DATA  = 3   # C

    period_to_col = {p: FIRST_DATA + i for i, p in enumerate(ordered_periods)}
    total_cols = FIRST_DATA + len(ordered_periods) - 1

    # -----------------------------------------------------------------------
    # 2. Set column widths
    # -----------------------------------------------------------------------
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 38
    for i, p in enumerate(ordered_periods):
        col_letter = get_column_letter(FIRST_DATA + i)
        ws.column_dimensions[col_letter].width = 11 if _is_annual(p) else 9

    # -----------------------------------------------------------------------
    # 3. ROW 1 – Company header
    # -----------------------------------------------------------------------
    ws.row_dimensions[1].height = 28
    cell = ws.cell(row=1, column=LABEL_COL, value=company_name)
    cell.font      = _font(bold=True, color=WHITE, size=14)
    cell.fill      = _fill(COL_HEADER_FILL)
    cell.alignment = _center()
    # Merge B1 across all data columns
    ws.merge_cells(start_row=1, start_column=LABEL_COL,
                   end_row=1,   end_column=total_cols)
    # Fill A1
    ws.cell(row=1, column=1).fill = _fill(COL_HEADER_FILL)

    # -----------------------------------------------------------------------
    # 4. ROW 2 – Period headers
    # -----------------------------------------------------------------------
    ws.row_dimensions[2].height = 22
    ws.cell(row=2, column=1, value='Type').font   = _font(bold=True)
    ws.cell(row=2, column=1).fill                  = _fill(COL_HEADER_FILL)
    ws.cell(row=2, column=1).border                = _border()
    ws.cell(row=2, column=2, value='(Rs. Cr.)').font = _font(bold=True)
    ws.cell(row=2, column=2).fill                    = _fill(COL_HEADER_FILL)
    ws.cell(row=2, column=2).border                  = _border()

    for p, col in period_to_col.items():
        c = ws.cell(row=2, column=col, value=p)
        c.fill      = _fill(COL_HEADER_FILL)
        c.font      = _font(bold=True, size=9)
        c.alignment = _center()
        c.border    = _border()

    # -----------------------------------------------------------------------
    # 5. Data rows
    # -----------------------------------------------------------------------
    GROWTH_ROWS   = {4, 7, 10, 29, 32, 40, 50}
    PERCENT_ROWS  = {14,16,18,20,22,25,27,30,33,57,58,60,61,62,79}
    RATIO_PCT     = {60, 61, 62, 79}

    for row_idx, (label_key, metric, cat) in ROW_SPEC.items():
        if cat == 'HEADER':
            continue
        ws.row_dimensions[row_idx].height = 16
        
        # Determine base fill color for this row
        row_fill = GROWTH_FILL if row_idx in GROWTH_ROWS else WHITE

        # Category cell (col A)
        ca = ws.cell(row=row_idx, column=1, value=cat if cat not in ('BLANK','HEADER') else '')
        ca.fill      = _fill(row_fill)
        ca.font      = _font(color=BLACK, size=8)
        ca.border    = _border()
        ca.alignment = _center()

        # Label cell (col B)
        label_display = label_key if cat != 'HEADER' else ''
        cl = ws.cell(row=row_idx, column=2, value=label_display)
        cl.fill      = _fill(row_fill)
        cl.font      = _font(bold=(cat in ('PL','BS','CF','RATIO') or 'Margin' in label_key or 'Growth' in label_key))
        cl.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cl.border    = _border()

        # Data cells
        for p, col in period_to_col.items():
            col_let = get_column_letter(col)
            val = None
            formula = None
            is_ann = _is_annual(p)

            # Check if this annual column follows quarters that we can sum dynamically
            prev_q_cols = []
            if is_ann:
                m_ann = re.match(r'^FY(\d{2})$', p)
                if m_ann:
                    fy_num = int(m_ann.group(1))
                    for q in range(1, 5):
                        q_key = f'{q}QFY{fy_num:02d}'
                        if q_key in period_to_col:
                            prev_q_cols.append(get_column_letter(period_to_col[q_key]))

            # ---------------------------------------------------------------------------
            # Build formulas — every formula is the SAME for quarterly and annual
            # (annual PL/CF rows will have =SUM(Q1:Q4) added below, overriding)
            # ---------------------------------------------------------------------------
            if row_idx == 6:   # Net Sales = Gross Sales - Excise
                formula = f"={col_let}3-{col_let}5"
            elif row_idx == 9:  # Total Income = Net Sales + Other Op Income
                formula = f"={col_let}6+{col_let}8"
            elif row_idx in (14, 16, 18, 20, 22):  # % sales
                base_row = {14: 13, 16: 15, 18: 17, 20: 19, 22: 21}[row_idx]
                formula = f"={col_let}{base_row}/IF({col_let}$6=0,1,{col_let}$6)"
            elif row_idx == 23:  # COGS
                formula = f"={col_let}13+{col_let}15+{col_let}17"
            elif row_idx == 24:  # Gross Profit
                formula = f"={col_let}6-({col_let}13+{col_let}15)"
            elif row_idx == 25:  # GP Margin
                formula = f"={col_let}24/IF({col_let}$6=0,1,{col_let}$6)"
            elif row_idx == 26:  # Total Expenditure
                formula = f"={col_let}13+{col_let}15+{col_let}17+{col_let}19+{col_let}21"
            elif row_idx == 27:  # TE % sales
                formula = f"={col_let}26/IF({col_let}$6=0,1,{col_let}$6)"
            elif row_idx == 28:  # EBITDA = Total Income - Total Exp
                formula = f"={col_let}9-{col_let}26"
            elif row_idx == 30:  # EBITDA Margin
                formula = f"={col_let}28/IF({col_let}9=0,1,{col_let}9)"
            elif row_idx == 31:  # PBIT = EBITDA - Depreciation
                formula = f"={col_let}28-{col_let}38"
            elif row_idx == 33:  # PBIT Margin
                formula = f"={col_let}31/IF({col_let}9=0,1,{col_let}9)"
            elif row_idx == 37:  # PBDT = EBITDA - Interest + Other Income
                formula = f"={col_let}28-{col_let}36+{col_let}35"
            elif row_idx == 39:  # PBT = PBDT - Depreciation
                formula = f"={col_let}37-{col_let}38"
            elif row_idx == 44:  # PAT = PBT - Tax - Deferred Tax
                formula = f"={col_let}39-{col_let}41-{col_let}43"
            elif row_idx == 47:  # Net Profit after MI
                formula = f"={col_let}44-{col_let}45"
            elif row_idx == 49:  # Adjusted PAT
                formula = f"={col_let}47+{col_let}48"
            elif row_idx == 56:  # Book Value per share
                formula = f"=IF({col_let}63=0,0,({col_let}51+{col_let}52)/({col_let}51/IF({col_let}63=0,1,{col_let}63)))"
            elif row_idx == 57:  # EBITDA Margin ratio
                formula = f"={col_let}28/IF({col_let}9=0,1,{col_let}9)"
            elif row_idx == 58:  # Tax/PBT %
                formula = f"=({col_let}41+{col_let}43)/IF({col_let}39=0,1,{col_let}39)"
            elif row_idx == 59:  # EPS
                formula = f"=IF({col_let}63=0,0,{col_let}49/({col_let}51/IF({col_let}63=0,1,{col_let}63)))"
            elif row_idx == 60:  # D:E
                formula = f"={col_let}53/IF(({col_let}51+{col_let}52)=0,1,({col_let}51+{col_let}52))"
            elif row_idx == 61:  # ROE %
                formula = f"={col_let}49/IF(({col_let}51+{col_let}52)=0,1,({col_let}51+{col_let}52))"
            elif row_idx == 62:  # ROCE %
                formula = f"={col_let}31/IF(({col_let}51+{col_let}52+{col_let}53)=0,1,({col_let}51+{col_let}52+{col_let}53))"
            elif row_idx == 64:  # No. of Shares = Equity / Face Value
                formula = f"=IF({col_let}63=0,0,{col_let}51/{col_let}63)"
            elif row_idx == 72:  # Working Capital
                formula = None  # written from data below
            elif row_idx == 76:  # NOPAT = PBIT * (1 - tax rate)
                formula = f"={col_let}31*(1-IF({col_let}39=0,0,({col_let}41+{col_let}43)/{col_let}39))"
            elif row_idx == 77:  # Capital Employed
                formula = f"={col_let}51+{col_let}52+{col_let}53"
            elif row_idx == 78:  # Invested Capital
                formula = f"={col_let}77-{col_let}54"
            elif row_idx == 79:  # ROIC
                formula = f"={col_let}76/IF({col_let}78=0,1,{col_let}78)"
            # ---------------------------------------------------------------------------
            # Growth rows — use same-row prior-period column reference for both Q and Annual
            # ---------------------------------------------------------------------------
            elif row_idx in (4, 7, 10, 29, 32, 40, 50):
                base_rows = {4:3, 7:6, 10:9, 29:28, 32:31, 40:39, 50:49}
                base_row = base_rows[row_idx]
                # Find the previous same-type period column
                prev_col_let = _get_prev_col(p, period_to_col, ordered_periods)
                if prev_col_let:
                    formula = f"={col_let}{base_row}/IF({prev_col_let}{base_row}=0,1,{prev_col_let}{base_row})-1"
                else:
                    formula = None
            # ---------------------------------------------------------------------------
            # Annual PL/CF rows: SUM of 4 quarters
            # ---------------------------------------------------------------------------
            elif is_ann and len(prev_q_cols) == 4 and cat in ('PL', 'CF') and metric not in (
                    'equity', 'reserves', 'debt', 'cash', 'face_value', 'net_block',
                    'working_capital', 'num_shares'):
                formula = f"=SUM({prev_q_cols[0]}{row_idx}:{prev_q_cols[-1]}{row_idx})"
            else:
                if metric and p in data:
                    val = _get_derived(data[p], metric, row_idx, data, p)

            c = ws.cell(row=row_idx, column=col)
            c.fill   = _fill(row_fill)
            c.border = _border()
            c.alignment = _right()

            if formula:
                c.value = formula
            elif val is not None:
                c.value = val

            if formula or val is not None:
                if row_idx in GROWTH_ROWS or row_idx in PERCENT_ROWS:
                    c.number_format = '0.0%'
                    c.font = _font(bold=is_ann, size=9)
                else:
                    c.number_format = '#,##0.00'
                    c.font = _font(bold=is_ann, size=9)
            else:
                c.font = _font(size=9)

    # -----------------------------------------------------------------------
    # 6. Freeze panes at C3
    # -----------------------------------------------------------------------
    ws.freeze_panes = 'C3'

    # -----------------------------------------------------------------------
    # 7. Autofilter on header row
    # -----------------------------------------------------------------------
    ws.auto_filter.ref = f'B2:{get_column_letter(total_cols)}2'

    wb.save(output_path)
    return output_path

# ---------------------------------------------------------------------------
# Helper: find previous same-type period column letter
# ---------------------------------------------------------------------------

def _get_prev_col(current_period: str, period_to_col: dict, ordered_periods: list) -> str:
    """
    Return the column letter of the previous same-type period.
    For quarterly: prior year same quarter (1QFY18 → 1QFY17).
    For annual: prior year FY (FY18 → FY17).
    Returns None if not found.
    """
    qm = re.match(r'^(\d)QFY(\d{2})$', current_period)
    am = re.match(r'^FY(\d{2})$', current_period)
    if qm:
        prev = f'{qm.group(1)}QFY{int(qm.group(2))-1:02d}'
    elif am:
        prev = f'FY{int(am.group(1))-1:02d}'
    else:
        return None
    if prev in period_to_col:
        return get_column_letter(period_to_col[prev])
    return None



def _get_derived(period_data: dict, metric: str, row_idx: int,
                 all_data: dict, current_period: str):
    """
    Compute / retrieve a metric for a single period.
    Returns float or None.
    """
    d = period_data
    g = d.get   # shorthand

    if metric == 'gross_profit':
        ns = g('net_sales') or g('total_income') or 0
        rm = g('raw_material') or 0
        sa = g('stock_adj') or 0
        return ns - rm - sa if ns else None

    if metric == 'gp_margin':
        gp = g('net_sales', 0) - (g('raw_material') or 0) - (g('stock_adj') or 0)
        ns = g('net_sales') or g('total_income')
        return gp / ns if ns else None

    if metric == 'rm_pct':
        ns = g('net_sales') or g('total_income')
        return (g('raw_material') or 0) / ns if ns else None

    if metric == 'sa_pct':
        ns = g('net_sales') or g('total_income')
        return (g('stock_adj') or 0) / ns if ns else None

    if metric == 'pfg_pct':
        ns = g('net_sales') or g('total_income')
        return (g('purchase_fg') or 0) / ns if ns else None

    if metric == 'emp_pct':
        ns = g('net_sales') or g('total_income')
        return (g('employee_exp') or 0) / ns if ns else None

    if metric == 'oe_pct':
        ns = g('net_sales') or g('total_income')
        return ((g('other_exp') or 0) + (g('other_exp2') or 0)) / ns if ns else None

    if metric == 'te_pct':
        ns = g('net_sales') or g('total_income')
        return (g('total_expenditure') or 0) / ns if ns else None

    if metric == 'tax_current':
        return d.get('tax_current') or d.get('current_tax') or d.get('tax') or d.get('total_tax')

    if metric == 'tax_pbt_pct':
        pbt = g('pbt')
        tax = g('total_tax') or g('tax') or 0
        return tax / pbt if pbt and pbt != 0 else None

    if metric == 'eps':
        pat  = g('pat') or g('net_profit_after_tax') or g('pat_owners')
        fv   = g('face_value') or 2
        eq   = g('equity')  # Share capital in Cr.
        if pat and eq and eq != 0:
            shares_cr = eq / fv   # number of shares in crores
            return pat / shares_cr if shares_cr != 0 else None
        return None

    if metric == 'book_value_ps':
        eq  = g('equity') or 0
        res = g('reserves') or 0
        fv  = g('face_value') or 2
        if eq != 0:
            shares_cr = eq / fv
            return (eq + res) / shares_cr if shares_cr else None
        return None

    if metric == 'num_shares':
        eq = g('equity')
        fv = g('face_value') or 2
        return eq / fv if eq else None

    if metric == 'nopat':
        pbit = g('pbit') or 0
        pbt  = g('pbt') or 0
        tax  = g('total_tax') or g('tax') or 0
        tax_rate = tax / pbt if pbt and pbt != 0 else 0
        return pbit * (1 - tax_rate)

    if metric == 'cap_employed':
        eq   = g('equity') or 0
        res  = g('reserves') or 0
        debt = g('debt') or 0
        return eq + res + debt

    if metric == 'invested_cap':
        cap  = (g('equity') or 0) + (g('reserves') or 0) + (g('debt') or 0)
        cash = g('cash') or 0
        return cap - cash

    if metric == 'roic':
        nopat = _get_derived(d, 'nopat', 0, all_data, current_period)
        ic    = _get_derived(d, 'invested_cap', 0, all_data, current_period)
        return nopat / ic if ic and ic != 0 else None

    # Growth metrics (YOY)
    GROWTH_MAP = {
        'gs_growth':    'gross_sales',
        'ns_growth':    'net_sales',
        'ti_growth':    'total_income',
        'ebitda_growth':'ebitda',
        'pbit_growth':  'pbit',
        'pbt_growth':   'pbt',
        'pat_growth':   'pat',
    }
    if metric in GROWTH_MAP:
        base_metric = GROWTH_MAP[metric]
        curr_val = d.get(base_metric)
        prev_period = _get_prev_period(current_period, all_data)
        if prev_period and curr_val:
            prev_val = all_data[prev_period].get(base_metric)
            if prev_val and prev_val != 0:
                return (curr_val - prev_val) / abs(prev_val)
        return None

    # Default: just return raw value
    return d.get(metric)


def _get_prev_period(current: str, all_data: dict):
    """Find the same-type period from 1 year earlier."""
    qm = re.match(r'^(\d)QFY(\d{2})$', current)
    am = re.match(r'^FY(\d{2})$', current)
    if qm:
        q, fy = int(qm.group(1)), int(qm.group(2))
        prev = f'{q}QFY{fy-1:02d}'
        return prev if prev in all_data else None
    if am:
        fy = int(am.group(1))
        prev = f'FY{fy-1:02d}'
        return prev if prev in all_data else None
    return None


# ---------------------------------------------------------------------------
# CLI test
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from parser import parse_all
    csv_dir = sys.argv[1] if len(sys.argv) > 1 else '..'
    out     = sys.argv[2] if len(sys.argv) > 2 else '/tmp/test_output.xlsx'
    data    = parse_all(csv_dir)
    write_excel(data, 'Supreme Industries Ltd', out)
    print(f'Written to {out}')
