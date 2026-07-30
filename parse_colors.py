import openpyxl
from openpyxl.styles import colors

def get_color(cell):
    fill = cell.fill
    if fill.fgColor and hasattr(fill.fgColor, 'rgb'):
        return fill.fgColor.rgb
    return None

def get_font_color(cell):
    font = cell.font
    if font.color and hasattr(font.color, 'rgb'):
        return font.color.rgb
    return None

wb = openpyxl.load_workbook('Supreme Ind_Format.xlsx', data_only=True)
ws = wb.active

print("--- EXCEL METADATA ---")
print("Show grid lines:", ws.sheet_view.showGridLines)

print("\n--- SAMPLE CELLS ---")
samples = [
    ("Company Header (B1)", ws['B1']),
    ("Type Header (A2)", ws['A2']),
    ("Q1 Period Header (C2)", ws['C2']),
    ("Gross Sales Label (B3)", ws['B3']),
    ("Gross Sales Data (C3)", ws['C3']),
    ("Growth % YOY Label (B4)", ws['B4']),
    ("Growth % YOY Data (C4)", ws['C4']),
    ("Raw Material Label (B13)", ws['B13']),
    ("PBIDT (EBITDA) Label (B28)", ws['B28']),
    ("Equity Label (B51)", ws['B51']),
    ("Blank Category (A11)", ws['A11']),
    ("Category A3", ws['A3']),
    ("Category A51", ws['A51']),
    ("Category A73", ws['A73'])
]

for name, cell in samples:
    fill_c = get_color(cell)
    font_c = get_font_color(cell)
    border_left = cell.border.left.style if cell.border.left else "none"
    border_bottom = cell.border.bottom.style if cell.border.bottom else "none"
    
    print(f"{name}:")
    print(f"  Fill: {fill_c}")
    print(f"  Font: {font_c} | Bold: {cell.font.bold}")
    print(f"  Borders: L={border_left}, B={border_bottom}")

