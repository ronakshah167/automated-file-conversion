import openpyxl

wb = openpyxl.load_workbook('Supreme Ind_Format.xlsx')
ws = wb.active

samples = [
    ("B1 (Company)", ws['B1']),
    ("C2 (Period)", ws['C2']),
    ("B3 (Gross Sales)", ws['B3']),
    ("B4 (Growth)", ws['B4'])
]

for name, cell in samples:
    fill = cell.fill
    if fill.patternType is not None and fill.fgColor.type == 'indexed':
        print(f"{name} indexed color: {fill.fgColor.indexed}")

